from fastapi import APIRouter, Request, Depends, Query
from fastapi.responses import RedirectResponse, HTMLResponse, Response
from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import Optional
import io
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
from app.database import get_db
from app.models.detection import Detection
from app.models.user import User
from app.routers.auth import get_current_user
from app.jinja_setup import templates

router = APIRouter(prefix="/reports", tags=["reports"])

def get_filtered_detections(
    db: Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None
):
    q = db.query(Detection).join(User, Detection.user_id == User.id, isouter=True)
    if start_date:
        sd = datetime.strptime(start_date, "%Y-%m-%d").date()
        q = q.filter(Detection.created_at >= sd)
    if end_date:
        ed = datetime.strptime(end_date, "%Y-%m-%d").date()
        q = q.filter(Detection.created_at <= ed + __import__("datetime").timedelta(days=1))
    return q.order_by(Detection.created_at.desc()).all()

def parse_objects(detection):
    if detection.detected_objects:
        try:
            return json.loads(detection.detected_objects)
        except (json.JSONDecodeError, TypeError):
            return []
    return []

def objects_text(objects):
    return ", ".join([o.get("label", "?") for o in objects])

@router.get("", response_class=HTMLResponse)
async def reports_page(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user)
):
    if not user:
        return RedirectResponse(url="/auth/login", status_code=302)
    detections = db.query(Detection).order_by(Detection.created_at.desc()).all()
    users = db.query(User).all()
    return templates.TemplateResponse(request, "reports/index.html", {
        "user": user,
        "detections": detections,
        "users": users
    })

@router.get("/export/excel")
async def export_excel(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    if not user:
        return RedirectResponse(url="/auth/login", status_code=302)

    detections = get_filtered_detections(db, start_date, end_date)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Deteksi"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["ID", "User", "Objek Terdeteksi", "Jumlah Objek", "Tanggal"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for i, d in enumerate(detections, 2):
        objs = parse_objects(d)
        row_data = [
            d.id,
            d.user.full_name if d.user else d.user.username if d.user else "-",
            objects_text(objs),
            len(objs),
            d.created_at.strftime("%d/%m/%Y %H:%M")
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"laporan_deteksi_{date.today().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=output.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@router.get("/export/pdf")
async def export_pdf(
    request: Request,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    if not user:
        return RedirectResponse(url="/auth/login", status_code=302)

    detections = get_filtered_detections(db, start_date, end_date)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "LAPORAN DETEKSI OBJEK", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 7, f"Periode: {start_date or 'Awal'} - {end_date or 'Sekarang'}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 7, f"Total Data: {len(detections)} deteksi", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

    col_w = [10, 35, 90, 20, 45]
    headers = ["ID", "User", "Objek", "Jml", "Tanggal"]

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    for i, h in enumerate(headers):
        pdf.cell(col_w[i], 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "", 8)
    for d in detections:
        objs = parse_objects(d)
        row = [
            str(d.id),
            d.user.full_name if d.user else d.user.username if d.user else "-",
            objects_text(objs)[:60],
            str(len(objs)),
            d.created_at.strftime("%d/%m/%Y %H:%M")
        ]
        for i, val in enumerate(row):
            pdf.cell(col_w[i], 7, val, border=1, align="C" if i in [0, 3] else "L")
        pdf.ln()

    output = io.BytesIO()
    pdf.output(output)
    output.seek(0)

    filename = f"laporan_deteksi_{date.today().strftime('%Y%m%d')}.pdf"
    return Response(
        content=output.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
